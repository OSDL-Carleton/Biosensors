"""
characterization.py - Complete Device Characterization Module
Usage: characterize_device(data_by_dac1, mode, ADC, save_path, params_entries)
"""

import numpy as np
import matplotlib.pyplot as plt
from scipy import signal, stats
import time
import os
import datetime
import openpyxl
from matplotlib.backends.backend_pdf import PdfPages
import tkinter.messagebox as msgbox


class DeviceCharacterizer:
    """Complete device characterization for transistors"""

    def __init__(self, save_path, chip_name, trial_name):
        self.baseline_noise_rms = None
        self.save_path = save_path
        self.chip_name = chip_name
        self.trial_name = trial_name
        self.results = {}

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        self.char_folder = os.path.join(save_path, f"Characterization_{timestamp}")
        os.makedirs(self.char_folder, exist_ok=True)

        print(f"Characterization results will be saved to: {self.char_folder}")

    def measure_baseline_noise(self, ADC, duration=5.0):
        """Measure baseline noise for SNR calculations"""
        print(f"\n=== BASELINE NOISE MEASUREMENT ===")
        print(f"Measuring baseline noise for {duration} seconds...")
        print("Ensure all inputs are grounded/disconnected")

        ready = msgbox.askyesno(
            "Baseline Noise Measurement",
            f"Ready to measure baseline noise?\n\n"
            f"Before proceeding, please ensure:\n"
            f"• Device is disconnected from all probes\n"
            f"• All ADC inputs are connected to ground\n"
            f"• All probe wires are grounded\n\n"
            f"Measurement will take {duration} seconds.\n\n"
            f"Click 'Yes' when ready to start.",
        )

        if not ready:
            print("Baseline noise measurement cancelled by user")
            return None, None

        noise_readings = []
        start_time = time.time()

        print(f"Starting baseline noise measurement...")
        while time.time() - start_time < duration:
            ADC_Value = ADC.ADS1256_GetAll()
            reading = ADC_Value[4] * 5.0 / 0x7FFFFF
            noise_readings.append(reading)
            time.sleep(0.01)

        self.baseline_noise_rms = np.std(noise_readings)
        print(f"Baseline noise: {self.baseline_noise_rms*1e6:.2f} µV RMS")

        msgbox.showinfo(
            "Baseline Noise Complete",
            f"Baseline noise measurement completed!\n\n"
            f"Measured noise: {self.baseline_noise_rms*1e6:.2f} µV RMS\n\n"
            f"You can now reconnect your device for characterization.",
        )

        self.save_excel_data(
            "baseline_noise",
            ["Time (s)", "Noise (V)"],
            [np.arange(len(noise_readings)) * 0.01, noise_readings],
        )

        return self.baseline_noise_rms, noise_readings

    def analyze_channel_resistance(self, vds_data, ids_data, gate_voltage):
        """Calculate and plot channel resistance (Ron)"""
        print(f"\n=== CHANNEL RESISTANCE (Ron) ANALYSIS ===")

        sorted_indices = np.argsort(vds_data)
        vds_sorted = np.array(vds_data)[sorted_indices]
        ids_sorted = np.array(ids_data)[sorted_indices]

        linear_region_end = int(0.2 * len(vds_sorted))
        if linear_region_end < 3:
            linear_region_end = min(5, len(vds_sorted))

        vds_linear = vds_sorted[:linear_region_end]
        ids_linear = ids_sorted[:linear_region_end]

        if len(vds_linear) > 1:
            slope, intercept = np.polyfit(vds_linear, ids_linear, 1)
            ron = 1.0 / slope if slope != 0 else float("inf")
            r_squared = self.calculate_r_squared(
                ids_linear, slope * vds_linear + intercept
            )
        else:
            ron = float("inf")
            r_squared = 0

        ron = abs(ron)

        print(f"Channel Resistance (Ron): {ron:.2f} Ω")
        print(f"Linear region R²: {r_squared:.4f}")

        fig, ax = plt.subplots(figsize=(10, 6))
        ax.plot(vds_sorted, ids_sorted * 1e6, "b-", linewidth=2, label="Output Curve")
        ax.plot(vds_linear, ids_linear * 1e6, "ro", markersize=6, label="Linear Region")

        if len(vds_linear) > 1:
            fit_line = slope * vds_linear + intercept
            ax.plot(
                vds_linear,
                fit_line * 1e6,
                "r--",
                linewidth=2,
                label=f"Linear Fit (Ron = {ron:.1f} Ω)",
            )

        ax.set_xlabel("VDS (V)")
        ax.set_ylabel("IDS (µA)")
        ax.set_title(f"Channel Resistance Analysis (Gate = {gate_voltage}V)")
        ax.legend()
        ax.grid(True, alpha=0.3)

        ax.text(
            0.02,
            0.98,
            f"Ron = {ron:.2f} Ω\nR² = {r_squared:.4f}",
            transform=ax.transAxes,
            fontsize=12,
            fontweight="bold",
            bbox=dict(boxstyle="round,pad=0.3", facecolor="lightblue", alpha=0.7),
            verticalalignment="top",
        )

        self.save_plot(fig, f"ron_analysis_gate_{gate_voltage}V")

        self.save_excel_data(
            f"ron_gate_{gate_voltage}V",
            ["VDS (V)", "IDS (A)", "Linear_Region", "Fit_Line (A)"],
            [
                vds_sorted,
                ids_sorted,
                np.concatenate(
                    [ids_linear, [np.nan] * (len(vds_sorted) - len(ids_linear))]
                ),
                np.concatenate(
                    [
                        slope * vds_linear + intercept,
                        [np.nan] * (len(vds_sorted) - len(vds_linear)),
                    ]
                ),
            ],
        )

        return {
            "ron": ron,
            "r_squared": r_squared,
            "linear_region_end": linear_region_end,
        }

    def analyze_output_conductance(self, vds_data, ids_data, gate_voltage):
        """Calculate and plot output conductance (gds)"""
        print(f"\n=== OUTPUT CONDUCTANCE (gds) ANALYSIS ===")

        sorted_indices = np.argsort(vds_data)
        vds_sorted = np.array(vds_data)[sorted_indices]
        ids_sorted = np.array(ids_data)[sorted_indices]

        sat_region_start = int(0.7 * len(vds_sorted))
        vds_sat = vds_sorted[sat_region_start:]
        ids_sat = ids_sorted[sat_region_start:]

        if len(vds_sat) > 1:

            gds_array = np.gradient(ids_sat, vds_sat)
            gds = np.mean(gds_array)
        else:
            gds = 0
            gds_array = [0]

        gds = abs(gds)

        print(f"Output Conductance (gds): {gds*1e6:.2f} µS")
        print(f"Saturation region: VDS > {vds_sorted[sat_region_start]:.2f} V")

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))

        ax1.plot(vds_sorted, ids_sorted * 1e6, "b-", linewidth=2, label="Output Curve")
        ax1.plot(vds_sat, ids_sat * 1e6, "go", markersize=4, label="Saturation Region")
        ax1.set_xlabel("VDS (V)")
        ax1.set_ylabel("IDS (µA)")
        ax1.set_title(f"Output Conductance Analysis (Gate = {gate_voltage}V)")
        ax1.legend()
        ax1.grid(True, alpha=0.3)

        if len(vds_sat) > 1:
            ax2.plot(
                vds_sat, gds_array * 1e6, "r-", linewidth=2, marker="o", markersize=4
            )
            ax2.axhline(
                gds * 1e6,
                color="orange",
                linestyle="--",
                linewidth=2,
                label=f"Average gds = {gds*1e6:.2f} µS",
            )

        ax2.set_xlabel("VDS (V)")
        ax2.set_ylabel("gds (µS)")
        ax2.set_title("Output Conductance vs VDS")
        ax2.legend()
        ax2.grid(True, alpha=0.3)

        self.save_plot(fig, f"gds_analysis_gate_{gate_voltage}V")

        self.save_excel_data(
            f"gds_gate_{gate_voltage}V",
            ["VDS (V)", "IDS (A)", "Saturation_Region", "gds (S)"],
            [
                vds_sorted,
                ids_sorted,
                np.concatenate([[np.nan] * sat_region_start, ids_sat]),
                np.concatenate([[np.nan] * sat_region_start, gds_array]),
            ],
        )

        return {"gds": gds, "sat_region_start": sat_region_start}

    def analyze_snr(self, signal_data, gate_voltage):
        """Calculate and plot SNR metrics"""
        if self.baseline_noise_rms is None:
            print("ERROR: Must measure baseline noise first")
            return None

        print(f"\n=== SNR ANALYSIS ===")

        signal_rms = np.sqrt(np.mean(np.array(signal_data) ** 2))
        snr_linear = signal_rms / self.baseline_noise_rms
        snr_db = 20 * np.log10(snr_linear)
        noise_floor = 3 * self.baseline_noise_rms
        current_detection_limit = noise_floor
        max_signal = np.max(np.abs(signal_data))
        dynamic_range_db = 20 * np.log10(max_signal / noise_floor)

        print(f"SNR: {snr_db:.1f} dB")
        print(f"Current Detection Limit: {current_detection_limit*1e9:.2f} nA")
        print(f"Dynamic Range: {dynamic_range_db:.1f} dB")
        print(f"Noise Floor: {noise_floor*1e12:.2f} pA")

        fig, ax = plt.subplots(figsize=(10, 6))

        ax.plot(signal_data * 1e6, "b-", linewidth=1.5, label="Signal")

        noise_floor_ua = noise_floor * 1e6
        ax.axhline(
            noise_floor_ua,
            color="r",
            linestyle="--",
            linewidth=2,
            label=f"Noise Floor (+{noise_floor_ua:.2f} µA)",
        )
        ax.axhline(
            -noise_floor_ua,
            color="r",
            linestyle="--",
            linewidth=2,
            label=f"Noise Floor (-{noise_floor_ua:.2f} µA)",
        )
        ax.fill_between(
            range(len(signal_data)),
            -noise_floor_ua,
            noise_floor_ua,
            alpha=0.2,
            color="red",
            label="Unreliable Region",
        )

        ax.set_xlabel("Measurement Point")
        ax.set_ylabel("Current (µA)")
        ax.set_title(f"SNR Analysis (Gate = {gate_voltage}V)")
        ax.legend()
        ax.grid(True, alpha=0.3)

        quality = "EXCELLENT" if snr_db > 40 else "GOOD" if snr_db > 20 else "POOR"
        quality_color = "green" if snr_db > 40 else "orange" if snr_db > 20 else "red"

        ax.text(
            0.02,
            0.98,
            f"SNR: {snr_db:.1f} dB\n{quality}\nDetection Limit: {current_detection_limit*1e9:.2f} nA",
            transform=ax.transAxes,
            fontsize=12,
            fontweight="bold",
            bbox=dict(boxstyle="round,pad=0.3", facecolor=quality_color, alpha=0.7),
            verticalalignment="top",
        )

        self.save_plot(fig, f"snr_analysis_gate_{gate_voltage}V")

        measurement_points = np.arange(len(signal_data))
        noise_floor_array = np.full_like(signal_data, noise_floor)

        self.save_excel_data(
            f"snr_gate_{gate_voltage}V",
            ["Measurement_Point", "Signal (A)", "Noise_Floor (A)", "SNR_dB"],
            [
                measurement_points,
                signal_data,
                noise_floor_array,
                [snr_db] * len(signal_data),
            ],
        )

        return {
            "snr_db": snr_db,
            "current_detection_limit": current_detection_limit,
            "dynamic_range_db": dynamic_range_db,
            "noise_floor": noise_floor,
        }

    def analyze_drift(self, time_data, current_data, gate_voltage):
        """Calculate and plot drift rate"""
        if len(time_data) != len(current_data) or len(time_data) < 2:
            print("ERROR: Need time series data for drift calculation")
            return None

        print(f"\n=== DRIFT ANALYSIS ===")

        drift_rate, offset = np.polyfit(time_data, current_data, 1)
        mean_current = np.mean(current_data)
        drift_rate_percent_per_hour = (
            (drift_rate * 3600 / mean_current * 100) if mean_current != 0 else 0
        )
        total_drift = drift_rate * (time_data[-1] - time_data[0])
        total_drift_percent = (
            (total_drift / mean_current * 100) if mean_current != 0 else 0
        )

        print(f"Drift Rate: {drift_rate*1e12:.2f} pA/s")
        print(f"Drift Rate: {drift_rate_percent_per_hour:.4f} %/hour")
        print(
            f"Total Drift: {total_drift_percent:.4f} % over {time_data[-1] - time_data[0]:.1f} s"
        )

        fig, ax = plt.subplots(figsize=(10, 6))

        ax.plot(
            time_data, current_data * 1e6, "b-", linewidth=2, label="Measured Current"
        )

        time_fit = np.array([min(time_data), max(time_data)])
        current_fit = drift_rate * time_fit + offset
        ax.plot(
            time_fit,
            current_fit * 1e6,
            "r--",
            linewidth=2,
            label=f"Drift Trend ({drift_rate_percent_per_hour:.4f} %/hr)",
        )

        ax.set_xlabel("Time (s)")
        ax.set_ylabel("Current (µA)")
        ax.set_title(f"Drift Analysis (Gate = {gate_voltage}V)")
        ax.legend()
        ax.grid(True, alpha=0.3)

        ax.text(
            0.02,
            0.98,
            f"Drift: {drift_rate_percent_per_hour:.4f} %/hr\n"
            f"Total: {total_drift_percent:.4f} %\n"
            f"Duration: {time_data[-1] - time_data[0]:.1f} s",
            transform=ax.transAxes,
            fontsize=10,
            fontweight="bold",
            bbox=dict(boxstyle="round,pad=0.3", facecolor="lightyellow", alpha=0.7),
            verticalalignment="top",
        )

        self.save_plot(fig, f"drift_analysis_gate_{gate_voltage}V")

        self.save_excel_data(
            f"drift_gate_{gate_voltage}V",
            ["Time (s)", "Current (A)", "Drift_Fit (A)"],
            [time_data, current_data, drift_rate * time_data + offset],
        )

        return {
            "drift_rate_percent_per_hour": drift_rate_percent_per_hour,
            "total_drift_percent": total_drift_percent,
            "drift_rate_A_per_s": drift_rate,
        }

    def analyze_1f_noise(self, current_data, gate_voltage, sampling_rate=100):
        """Analyze and plot 1/f noise"""
        if len(current_data) < 100:
            print("ERROR: Need at least 100 data points for 1/f noise analysis")
            return None

        print(f"\n=== 1/f NOISE ANALYSIS ===")

        freqs, psd = signal.welch(
            current_data, fs=sampling_rate, nperseg=len(current_data) // 4
        )

        low_freq_mask = (freqs > 0.1) & (freqs < 10)

        if np.sum(low_freq_mask) > 5:
            log_freqs = np.log10(freqs[low_freq_mask])
            log_psd = np.log10(psd[low_freq_mask])

            slope, intercept = np.polyfit(log_freqs, log_psd, 1)
            noise_1f_coefficient = 10**intercept
            noise_1f_exponent = -slope
            noise_at_1hz = noise_1f_coefficient

            print(f"1/f Noise Coefficient: {noise_1f_coefficient:.2e} A²·Hz")
            print(f"1/f Noise Exponent: {noise_1f_exponent:.2f}")
            print(f"Current Noise at 1 Hz: {np.sqrt(noise_at_1hz)*1e12:.2f} pA/√Hz")

            fig, ax = plt.subplots(figsize=(10, 6))

            ax.loglog(
                freqs, np.sqrt(psd) * 1e12, "b-", linewidth=2, label="Measured Noise"
            )

            f_fit = np.logspace(-1, 2, 100)
            psd_fit = noise_1f_coefficient / (f_fit**noise_1f_exponent)
            ax.loglog(
                f_fit,
                np.sqrt(psd_fit) * 1e12,
                "r--",
                linewidth=2,
                label=f"1/f^{noise_1f_exponent:.2f} fit",
            )

            ax.set_xlabel("Frequency (Hz)")
            ax.set_ylabel("Current Noise (pA/√Hz)")
            ax.set_title(f"1/f Noise Analysis (Gate = {gate_voltage}V)")
            ax.legend()
            ax.grid(True, alpha=0.3, which="both")

            ax.text(
                0.02,
                0.98,
                f"1/f Exponent: {noise_1f_exponent:.2f}\n"
                f"Noise @ 1Hz: {np.sqrt(noise_at_1hz)*1e12:.1f} pA/√Hz\n"
                f"Coefficient: {noise_1f_coefficient:.2e} A²·Hz",
                transform=ax.transAxes,
                fontsize=10,
                fontweight="bold",
                bbox=dict(boxstyle="round,pad=0.3", facecolor="lightcyan", alpha=0.7),
                verticalalignment="top",
            )

            self.save_plot(fig, f"1f_noise_analysis_gate_{gate_voltage}V")

            self.save_excel_data(
                f"1f_noise_gate_{gate_voltage}V",
                [
                    "Frequency (Hz)",
                    "PSD (A²/Hz)",
                    "Noise_Density (A/√Hz)",
                    "Fit (A/√Hz)",
                ],
                [
                    freqs,
                    psd,
                    np.sqrt(psd),
                    np.sqrt(noise_1f_coefficient / (freqs**noise_1f_exponent)),
                ],
            )

            return {
                "1f_coefficient": noise_1f_coefficient,
                "1f_exponent": noise_1f_exponent,
                "noise_at_1hz": noise_at_1hz,
            }
        else:
            print("ERROR: Insufficient data points in 1/f frequency range")
            return None

    def analyze_threshold_voltage(self, vgs_data, ids_data, drain_voltage):
        """Calculate and plot threshold voltage"""
        print(f"\n=== THRESHOLD VOLTAGE (Vth) ANALYSIS ===")

        sorted_indices = np.argsort(vgs_data)
        vgs_sorted = np.array(vgs_data)[sorted_indices]
        ids_sorted = np.array(ids_data)[sorted_indices]

        gm = np.abs(np.gradient(ids_sorted, vgs_sorted))
        max_gm_idx = np.argmax(gm)

        fit_range = min(5, len(vgs_sorted) // 4)
        start_idx = max(0, max_gm_idx - fit_range)
        end_idx = min(len(vgs_sorted), max_gm_idx + fit_range)

        vgs_fit = vgs_sorted[start_idx:end_idx]
        ids_fit = ids_sorted[start_idx:end_idx]

        if len(vgs_fit) > 1:
            slope, intercept = np.polyfit(vgs_fit, ids_fit, 1)
            vth = -intercept / slope if slope != 0 else 0
        else:
            vth = 0
            slope = 0

        print(f"Threshold Voltage (Vth): {vth:.3f} V")
        print(f"Transconductance region slope: {slope*1e6:.2f} µA/V")

        fig, ax = plt.subplots(figsize=(10, 6))

        ax.semilogy(
            vgs_sorted,
            np.abs(ids_sorted) * 1e6,
            "b-",
            linewidth=2,
            label="Transfer Curve",
        )
        ax.plot(
            vgs_fit,
            np.abs(ids_fit) * 1e6,
            "ro",
            markersize=6,
            label="Linear Extrapolation Region",
        )

        if len(vgs_fit) > 1:

            vgs_extrap = np.linspace(min(vgs_sorted), max(vgs_sorted), 100)
            ids_extrap = slope * vgs_extrap + intercept
            mask = ids_extrap > 0
            ax.semilogy(
                vgs_extrap[mask],
                ids_extrap[mask] * 1e6,
                "r--",
                linewidth=2,
                label=f"Linear Extrapolation",
            )

            ax.axvline(
                vth,
                color="orange",
                linestyle=":",
                linewidth=2,
                label=f"Vth = {vth:.3f} V",
            )

        ax.set_xlabel("VGS (V)")
        ax.set_ylabel("|IDS| (µA)")
        ax.set_title(f"Threshold Voltage Analysis (Drain = {drain_voltage}V)")
        ax.legend()
        ax.grid(True, alpha=0.3)

        ax.text(
            0.02,
            0.98,
            f"Vth = {vth:.3f} V",
            transform=ax.transAxes,
            fontsize=12,
            fontweight="bold",
            bbox=dict(boxstyle="round,pad=0.3", facecolor="lightgreen", alpha=0.7),
            verticalalignment="top",
        )

        self.save_plot(fig, f"vth_analysis_drain_{drain_voltage}V")

        extrapolation_line = slope * vgs_sorted + intercept
        self.save_excel_data(
            f"vth_drain_{drain_voltage}V",
            ["VGS (V)", "IDS (A)", "Extrapolation_Line (A)", "Vth (V)"],
            [vgs_sorted, ids_sorted, extrapolation_line, [vth] * len(vgs_sorted)],
        )

        return {"vth": vth, "slope": slope}

    def analyze_transconductance(self, vgs_data, ids_data, drain_voltage):
        """Calculate and plot transconductance"""
        print(f"\n=== TRANSCONDUCTANCE (gm) ANALYSIS ===")

        sorted_indices = np.argsort(vgs_data)
        vgs_sorted = np.array(vgs_data)[sorted_indices]
        ids_sorted = np.array(ids_data)[sorted_indices]

        gm = np.gradient(ids_sorted, vgs_sorted)
        gm_max = np.max(np.abs(gm))
        gm_max_idx = np.argmax(np.abs(gm))
        vgs_gm_max = vgs_sorted[gm_max_idx]

        print(f"Maximum Transconductance (gm): {gm_max*1e6:.2f} µS")
        print(f"gm_max occurs at VGS = {vgs_gm_max:.3f} V")

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))

        ax1.plot(
            vgs_sorted, ids_sorted * 1e6, "b-", linewidth=2, label="Transfer Curve"
        )
        ax1.axvline(
            vgs_gm_max,
            color="red",
            linestyle="--",
            alpha=0.7,
            label=f"Max gm at VGS = {vgs_gm_max:.3f} V",
        )
        ax1.set_xlabel("VGS (V)")
        ax1.set_ylabel("IDS (µA)")
        ax1.set_title(f"Transfer Characteristics (Drain = {drain_voltage}V)")
        ax1.legend()
        ax1.grid(True, alpha=0.3)

        ax2.plot(vgs_sorted, gm * 1e6, "g-", linewidth=2, label="Transconductance")
        ax2.plot(
            vgs_gm_max,
            gm_max * 1e6,
            "ro",
            markersize=8,
            label=f"Max gm = {gm_max*1e6:.2f} µS",
        )
        ax2.set_xlabel("VGS (V)")
        ax2.set_ylabel("gm (µS)")
        ax2.set_title("Transconductance vs Gate Voltage")
        ax2.legend()
        ax2.grid(True, alpha=0.3)

        self.save_plot(fig, f"gm_analysis_drain_{drain_voltage}V")

        self.save_excel_data(
            f"gm_drain_{drain_voltage}V",
            ["VGS (V)", "IDS (A)", "gm (S)", "gm_max (S)", "VGS_gm_max (V)"],
            [
                vgs_sorted,
                ids_sorted,
                gm,
                [gm_max] * len(vgs_sorted),
                [vgs_gm_max] * len(vgs_sorted),
            ],
        )

        return {"gm_max": gm_max, "vgs_gm_max": vgs_gm_max, "gm_curve": gm}

    def analyze_subthreshold_slope(self, vgs_data, ids_data, drain_voltage):
        """Calculate and plot subthreshold slope"""
        print(f"\n=== SUBTHRESHOLD SLOPE (SS) ANALYSIS ===")

        sorted_indices = np.argsort(vgs_data)
        vgs_sorted = np.array(vgs_data)[sorted_indices]
        ids_sorted = np.array(ids_data)[sorted_indices]

        log_ids = np.log10(np.abs(ids_sorted) + 1e-15)

        d_log_ids = np.gradient(log_ids, vgs_sorted)

        valid_indices = np.where(d_log_ids > 0)[0]
        if len(valid_indices) > 5:
            max_derivative_idx = valid_indices[np.argmax(d_log_ids[valid_indices])]
            max_derivative = d_log_ids[max_derivative_idx]
            ss = 1000 / max_derivative
            vgs_ss = vgs_sorted[max_derivative_idx]
        else:
            ss = float("inf")
            vgs_ss = 0
            max_derivative = 0

        print(f"Subthreshold Slope (SS): {ss:.1f} mV/decade")
        print(f"SS measured at VGS = {vgs_ss:.3f} V")

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))

        ax1.semilogy(
            vgs_sorted,
            np.abs(ids_sorted) * 1e6,
            "b-",
            linewidth=2,
            label="Transfer Curve",
        )
        if ss != float("inf"):
            ax1.axvline(
                vgs_ss,
                color="red",
                linestyle="--",
                alpha=0.7,
                label=f"SS measurement point",
            )
        ax1.set_xlabel("VGS (V)")
        ax1.set_ylabel("|IDS| (µA)")
        ax1.set_title(f"Transfer Curve (Log Scale) - Drain = {drain_voltage}V")
        ax1.legend()
        ax1.grid(True, alpha=0.3)

        ax2.plot(vgs_sorted, 1000 / d_log_ids, "r-", linewidth=2, label="Local SS")
        if ss != float("inf"):
            ax2.axhline(
                ss,
                color="orange",
                linestyle="--",
                linewidth=2,
                label=f"Min SS = {ss:.1f} mV/dec",
            )
            ax2.plot(vgs_ss, ss, "go", markersize=8, label="SS measurement point")
        ax2.set_xlabel("VGS (V)")
        ax2.set_ylabel("Subthreshold Slope (mV/decade)")
        ax2.set_title("Subthreshold Slope vs Gate Voltage")
        ax2.set_ylim(0, min(1000, np.percentile(1000 / d_log_ids[d_log_ids > 0], 95)))
        ax2.legend()
        ax2.grid(True, alpha=0.3)

        self.save_plot(fig, f"ss_analysis_drain_{drain_voltage}V")

        local_ss = 1000 / d_log_ids
        self.save_excel_data(
            f"ss_drain_{drain_voltage}V",
            ["VGS (V)", "IDS (A)", "Log_IDS", "Local_SS (mV/dec)", "Min_SS (mV/dec)"],
            [vgs_sorted, ids_sorted, log_ids, local_ss, [ss] * len(vgs_sorted)],
        )

        return {"ss": ss, "vgs_ss": vgs_ss}

    def analyze_on_off_ratio(self, vgs_data, ids_data, drain_voltage):
        """Calculate on/off ratio"""
        print(f"\n=== ON/OFF RATIO ANALYSIS ===")

        sorted_indices = np.argsort(vgs_data)
        vgs_sorted = np.array(vgs_data)[sorted_indices]
        ids_sorted = np.array(ids_data)[sorted_indices]

        i_on = np.max(np.abs(ids_sorted))
        i_off = (
            np.min(np.abs(ids_sorted[ids_sorted != 0]))
            if len(ids_sorted[ids_sorted != 0]) > 0
            else 1e-15
        )
        on_off_ratio = i_on / i_off

        i_on_idx = np.argmax(np.abs(ids_sorted))
        i_off_idx = np.argmin(np.abs(ids_sorted))
        vgs_ion = vgs_sorted[i_on_idx]
        vgs_ioff = vgs_sorted[i_off_idx]

        print(f"On Current (Ion): {i_on*1e6:.2f} µA at VGS = {vgs_ion:.3f} V")
        print(f"Off Current (Ioff): {i_off*1e9:.2f} nA at VGS = {vgs_ioff:.3f} V")
        print(f"On/Off Ratio: {on_off_ratio:.2e}")

        fig, ax = plt.subplots(figsize=(10, 6))

        ax.semilogy(
            vgs_sorted,
            np.abs(ids_sorted) * 1e6,
            "b-",
            linewidth=2,
            label="Transfer Curve",
        )
        ax.plot(
            vgs_ion, i_on * 1e6, "go", markersize=8, label=f"Ion = {i_on*1e6:.2f} µA"
        )
        ax.plot(
            vgs_ioff,
            i_off * 1e6,
            "ro",
            markersize=8,
            label=f"Ioff = {i_off*1e9:.2f} nA",
        )

        ax.set_xlabel("VGS (V)")
        ax.set_ylabel("|IDS| (µA)")
        ax.set_title(f"On/Off Ratio Analysis (Drain = {drain_voltage}V)")
        ax.legend()
        ax.grid(True, alpha=0.3)

        ax.text(
            0.02,
            0.98,
            f"Ion/Ioff = {on_off_ratio:.2e}",
            transform=ax.transAxes,
            fontsize=12,
            fontweight="bold",
            bbox=dict(boxstyle="round,pad=0.3", facecolor="lightpink", alpha=0.7),
            verticalalignment="top",
        )

        self.save_plot(fig, f"on_off_ratio_drain_{drain_voltage}V")

        self.save_excel_data(
            f"on_off_drain_{drain_voltage}V",
            ["VGS (V)", "IDS (A)", "Ion (A)", "Ioff (A)", "Ion_Ioff_Ratio"],
            [
                vgs_sorted,
                ids_sorted,
                [i_on] * len(vgs_sorted),
                [i_off] * len(vgs_sorted),
                [on_off_ratio] * len(vgs_sorted),
            ],
        )

        return {
            "on_off_ratio": on_off_ratio,
            "i_on": i_on,
            "i_off": i_off,
            "vgs_ion": vgs_ion,
            "vgs_ioff": vgs_ioff,
        }

    def calculate_r_squared(self, actual, predicted):
        """Calculate R-squared value"""
        ss_res = np.sum((actual - predicted) ** 2)
        ss_tot = np.sum((actual - np.mean(actual)) ** 2)
        return 1 - (ss_res / ss_tot) if ss_tot != 0 else 0

    def save_plot(self, fig, filename):
        """Save plot as PNG and PDF"""
        png_path = os.path.join(self.char_folder, f"{filename}.png")
        pdf_path = os.path.join(self.char_folder, f"{filename}.pdf")

        fig.savefig(png_path, dpi=300, bbox_inches="tight")
        fig.savefig(pdf_path, bbox_inches="tight")

        print(f"Plot saved: {filename}")
        plt.close(fig)

    def save_excel_data(self, filename, column_names, data_arrays):
        """Save data to Excel file"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Characterization Data"

        for col, header in enumerate(column_names, 1):
            worksheet.cell(row=1, column=col, value=header)

        max_length = max(len(arr) for arr in data_arrays)
        for row in range(max_length):
            for col, data_array in enumerate(data_arrays, 1):
                if row < len(data_array):
                    worksheet.cell(
                        row=row + 2, column=col, value=float(data_array[row])
                    )

        excel_path = os.path.join(self.char_folder, f"{filename}.xlsx")
        workbook.save(excel_path)
        print(f"Data saved: {filename}.xlsx")

    def create_summary_report(self):
        """Create a summary report of all characterization results"""
        summary_path = os.path.join(self.char_folder, "characterization_summary.txt")

        with open(summary_path, "w") as f:
            f.write(f"DEVICE CHARACTERIZATION SUMMARY\n")
            f.write(f"{'='*50}\n")
            f.write(f"Chip: {self.chip_name}\n")
            f.write(f"Trial: {self.trial_name}\n")
            f.write(
                f"Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
            )

            if self.baseline_noise_rms:
                f.write(f"Baseline Noise: {self.baseline_noise_rms*1e6:.2f} µV RMS\n\n")

            for condition, results in self.results.items():
                f.write(f"CONDITION: {condition}\n")
                f.write(f"{'-'*30}\n")
                for param, value in results.items():
                    if isinstance(value, dict):
                        for subparam, subvalue in value.items():
                            if isinstance(subvalue, (int, float)):
                                f.write(f"{subparam}: {subvalue}\n")
                    elif isinstance(value, (int, float)):
                        f.write(f"{param}: {value}\n")
                f.write("\n")

        print(f"Summary report saved: characterization_summary.txt")


def characterize_device(
    data_by_dac1,
    mode,
    ADC,
    save_path,
    params_entries,
    chip_name="Default_Chip",
    trial_name="Trial_1",
):
    """
    Main characterization function

    Args:
        data_by_dac1: Measurement data dictionary
        mode: "output" or "transfer"
        ADC: ADC object for baseline noise measurement
        save_path: Path to save results
        params_entries: Parameter entries from GUI
        chip_name: Name of the chip
        trial_name: Name of the trial
    """

    print(f"\n{'='*60}")
    print(f"STARTING {mode.upper()} MODE CHARACTERIZATION")
    print(f"{'='*60}")

    characterizer = DeviceCharacterizer(save_path, chip_name, trial_name)

    baseline_noise_rms, noise_data = characterizer.measure_baseline_noise(
        ADC, duration=5.0
    )

    if baseline_noise_rms is None:
        print("Characterization cancelled - baseline noise measurement required")
        return None

    for voltage_condition in data_by_dac1:
        print(f"\n{'='*50}")
        print(f"ANALYZING VOLTAGE CONDITION: {voltage_condition}V")
        print(f"{'='*50}")

        condition_results = {}

        if mode == "output":

            vds_data = np.array(data_by_dac1[voltage_condition]["vsd"])
            ids_data = np.array(data_by_dac1[voltage_condition]["isd"])

            condition_results["ron"] = characterizer.analyze_channel_resistance(
                vds_data, ids_data, voltage_condition
            )
            condition_results["gds"] = characterizer.analyze_output_conductance(
                vds_data, ids_data, voltage_condition
            )
            condition_results["snr"] = characterizer.analyze_snr(
                ids_data, voltage_condition
            )
            condition_results["drift"] = characterizer.analyze_drift(
                np.arange(len(ids_data)) * float(params_entries["step_interval"].get()),
                ids_data,
                voltage_condition,
            )

            if len(ids_data) > 100:
                condition_results["1f_noise"] = characterizer.analyze_1f_noise(
                    ids_data, voltage_condition
                )

        elif mode == "transfer":

            vgs_data = np.array(data_by_dac1[voltage_condition]["vsg"])
            ids_data = np.array(data_by_dac1[voltage_condition]["i_drain"])

            condition_results["vth"] = characterizer.analyze_threshold_voltage(
                vgs_data, ids_data, voltage_condition
            )
            condition_results["gm"] = characterizer.analyze_transconductance(
                vgs_data, ids_data, voltage_condition
            )
            condition_results["ss"] = characterizer.analyze_subthreshold_slope(
                vgs_data, ids_data, voltage_condition
            )
            condition_results["on_off"] = characterizer.analyze_on_off_ratio(
                vgs_data, ids_data, voltage_condition
            )

        characterizer.results[f"{mode}_{voltage_condition}V"] = condition_results

    characterizer.create_summary_report()

    print(f"\n{'='*60}")
    print(f"{mode.upper()} MODE CHARACTERIZATION COMPLETE")
    print(f"Results saved to: {characterizer.char_folder}")
    print(f"{'='*60}")

    return characterizer
