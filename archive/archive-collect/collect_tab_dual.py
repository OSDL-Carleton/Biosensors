import time
import datetime
import os
import RPi.GPIO as GPIO
import tkinter as tk
from tkinter import ttk
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from . import ADS1256
from . import DAC8532

GPIO.setwarnings(False)
GPIO.setmode(GPIO.BCM)

ADC = ADS1256.ADS1256()
DAC = DAC8532.DAC8532()
ADC.ADS1256_init()
DAC.DAC8532_Out_Voltage(0x30, 3)
DAC.DAC8532_Out_Voltage(0x34, 3)

collecting_data = False
voltages1, voltages2, voltages3 = [], [], []
voltages4, voltages5, voltages6 = [], [], []
vsd_values, i_source_values, i_gate_values, i_drain_values, isd_values = (
    [],
    [],
    [],
    [],
    [],
)
vsg_values = []
mode = "output"


def collect_tab(tab_collect, root):
    tab_collect.grid_rowconfigure(0, weight=1)
    tab_collect.grid_columnconfigure(0, weight=0, minsize=380)
    tab_collect.grid_columnconfigure(1, weight=1)

    container = tk.Frame(tab_collect, bg="lightgrey")
    container.grid(row=0, column=0, sticky="nswe", padx=10, pady=10)

    container.pack_propagate(False)

    canvas = tk.Canvas(container, borderwidth=0, background="lightgrey")
    vscroll = tk.Scrollbar(container, orient="vertical", command=canvas.yview, width=20)
    canvas.configure(yscrollcommand=vscroll.set)

    vscroll.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)

    scrollable_frame = tk.Frame(canvas, background="lightgrey")
    canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

    def on_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    def on_canvas_configure(event):
        canvas.itemconfig(canvas_window, width=canvas.winfo_width())

    scrollable_frame.bind("<Configure>", on_configure)
    canvas.bind("<Configure>", on_canvas_configure)

    def on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    canvas.bind("<MouseWheel>", on_mousewheel)
    canvas.bind("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
    canvas.bind("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))

    notebook_collect = ttk.Notebook(scrollable_frame)
    notebook_collect.pack(fill="both", expand=True)

    chip_name_entry, trial_name_entry, mode_var = create_data_setup_frame(notebook_collect)
    params_entries = create_params_frame(notebook_collect)
    create_console_log_frame(notebook_collect)
    create_controls_frame(
        notebook_collect, chip_name_entry, trial_name_entry, params_entries, mode_var, root
    )

    plot_frame_collect = tk.Frame(tab_collect)
    plot_frame_collect.grid(row=0, column=1, sticky="nswe")
    create_plot_frame(plot_frame_collect)


def create_controls_frame(
    parent, chip_name_entry, trial_name_entry, params_entries, mode_var, root
):
    frame_controls = tk.LabelFrame(
        parent, text="Controls", relief=tk.SUNKEN, borderwidth=2
    )
    frame_controls.pack(fill="x", padx=10, pady=5)

    def on_run():
        global collecting_data, mode
        mode = mode_var.get()
        sweep_start = float(params_entries["sweep_min"].get())
        sweep_end = float(params_entries["sweep_max"].get())
        sweep_step = float(params_entries["sweep_step"].get())
        interval = float(params_entries["step_interval"].get())
        duration = (
            float(params_entries["duration"].get())
            if "duration" in params_entries
            else 120
        )

        constant_voltage_values = [
            float(v.strip()) for v in params_entries["constant_voltage"].get().split(",")
        ]

        collecting_data = True
        perform_data_collection(
            sweep_start,
            sweep_end,
            sweep_step,
            interval,
            duration,
            constant_voltage_values,
            root,
        )

    def on_end():
        global collecting_data
        collecting_data = False

    def on_save():
        chip_name = chip_name_entry.get()
        trial_name = trial_name_entry.get()

        sweep_min = float(params_entries["sweep_min"].get())
        sweep_max = float(params_entries["sweep_max"].get())
        sweep_step = float(params_entries["sweep_step"].get())

        sweep_points = int((sweep_max - sweep_min) / sweep_step) + 1

        save_data(chip_name, trial_name, sweep_min, sweep_max, sweep_points)

    def on_reset():
        global voltages1, voltages2, voltages3, voltages4, voltages5, voltages6
        global vsd_values, i_source_values, i_gate_values, i_drain_values, isd_values, vsg_values
        voltages1, voltages2, voltages3 = [], [], []
        voltages4, voltages5, voltages6 = [], [], []
        vsd_values, i_source_values, i_gate_values, i_drain_values, isd_values = (
            [],
            [],
            [],
            [],
            [],
        )
        vsg_values = []
        reset_plot()
        reset_parameters(params_entries)

    run_button = tk.Button(frame_controls, text="Run", width=15, command=on_run)
    run_button.grid(row=0, column=0, padx=5, pady=5)

    end_button = tk.Button(frame_controls, text="End", width=15, command=on_end)
    end_button.grid(row=0, column=1, padx=5, pady=5)

    save_button = tk.Button(frame_controls, text="Save", width=15, command=on_save)
    save_button.grid(row=1, column=0, padx=5, pady=5)

    reset_button = tk.Button(frame_controls, text="Reset", width=15, command=on_reset)
    reset_button.grid(row=1, column=1, padx=5, pady=5)


plot_colors = ["blue", "green", "red", "orange", "purple", "brown", "cyan", "magenta"]

data_by_dac1 = {}


def perform_data_collection(
    sweep_start, sweep_end, sweep_step, interval, duration, constant_voltage_values, root
):
    global voltages1, voltages2, voltages3, voltages4, voltages5, voltages6
    global vsd_values, vsg_values, i_source_values, i_drain_values, isd_values, data_by_dac1
    global collecting_data, current_voltage, steps_taken, dac_index, mode

    data_by_dac1 = {
        v: {
            "voltages1": [],
            "voltages2": [],
            "voltages3": [],
            "voltages4": [],
            "voltages5": [],
            "voltages6": [],
            "vsd": [],
            "vsg": [],
            "i_source": [],
            "i_gate": [],
            "i_drain": [],
            "isd": [],
        }
        for v in constant_voltage_values
    }

    number_of_steps = int((sweep_end - sweep_start) / sweep_step) + 1

    dac_index = 0

    def collect_data():
        global collecting_data, current_voltage, steps_taken, dac_index

        if dac_index >= len(constant_voltage_values) or not collecting_data:
            collecting_data = False
            return

        constant_voltage = constant_voltage_values[dac_index]
        current_voltage = sweep_start
        steps_taken = 0
        end_time = time.time() + duration

        def collect_step():
            global current_voltage, steps_taken, dac_index

            if not collecting_data:
                return

            if steps_taken >= number_of_steps or time.time() >= end_time:
                dac_index += 1
                root.after(5000, collect_data)
                return

            if mode == "output":
                DAC.DAC8532_Out_Voltage(DAC8532.channel_B, constant_voltage)
                DAC.DAC8532_Out_Voltage(DAC8532.channel_A, current_voltage)
            else:
                DAC.DAC8532_Out_Voltage(DAC8532.channel_B, constant_voltage)
                DAC.DAC8532_Out_Voltage(DAC8532.channel_A, current_voltage)
                
            time.sleep(0.05)
            ADC_Value = ADC.ADS1256_GetAll()

            adc1 = ADC_Value[1] * 5.0 / 0x7FFFFF
            adc2 = ADC_Value[2] * 5.0 / 0x7FFFFF
            adc3 = ADC_Value[3] * 5.0 / 0x7FFFFF
            adc4 = ADC_Value[4] * 5.0 / 0x7FFFFF
            adc5 = ADC_Value[5] * 5.0 / 0x7FFFFF
            adc6 = ADC_Value[6] * 5.0 / 0x7FFFFF

            v1 = adc1
            v2 = adc2
            v3 = adc3
            v4 = adc4
            v5 = adc5
            v6 = adc6

            i_source = (adc4 - adc1) / 100.0
            
            if mode == "output":
                i_drain = (adc6 - adc3) / 100.0
                i_gate = (adc5 - adc2) / 1000.0
                vsd = adc4 - adc6
                vsg = adc4 - adc5
            else:
                i_drain = (adc5 - adc2) / 100.0
                i_gate = (adc6 - adc3) / 1000.0
                vsd = adc4 - adc5
                vsg = adc4 - adc6

            if mode == "output":
                isd = i_source - i_drain
                x_value = vsd
                y_value = isd
                x_label = "Vsd (V)"
                y_label = "Isd (A)"
            else:
                x_value = vsg
                y_value = i_drain
                x_label = "Vsg (V)"
                y_label = "Id (A)"

            voltages1.append(v1)
            voltages2.append(v2)
            voltages3.append(v3)
            voltages4.append(v4)
            voltages5.append(v5)
            voltages6.append(v6)
            i_source_values.append(i_source)
            i_gate_values.append(i_gate)
            i_drain_values.append(i_drain)
            vsd_values.append(vsd)
            vsg_values.append(vsg)

            if mode == "output":
                isd_values.append(isd)
                data_by_dac1[constant_voltage]["isd"].append(isd)

            data_by_dac1[constant_voltage]["voltages1"].append(v1)
            data_by_dac1[constant_voltage]["voltages2"].append(v2)
            data_by_dac1[constant_voltage]["voltages3"].append(v3)
            data_by_dac1[constant_voltage]["voltages4"].append(v4)
            data_by_dac1[constant_voltage]["voltages5"].append(v5)
            data_by_dac1[constant_voltage]["voltages6"].append(v6)
            data_by_dac1[constant_voltage]["vsd"].append(vsd)
            data_by_dac1[constant_voltage]["vsg"].append(vsg)
            data_by_dac1[constant_voltage]["i_source"].append(i_source)
            data_by_dac1[constant_voltage]["i_gate"].append(i_gate)
            data_by_dac1[constant_voltage]["i_drain"].append(i_drain)

            current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            update_plot(constant_voltage, x_value, y_value, x_label, y_label)

            if mode == "output":
                log_entry = (
                    f"Time: {current_time}\n"
                    f"ADC1 (Src-): {adc1:.5f} V, ADC2 (Drn-): {adc2:.5f} V, ADC3 (Gate-): {adc3:.5f} V\n"
                    f"ADC4 (Src+): {adc4:.5f} V, ADC5 (Drn+): {adc5:.5f} V, ADC6 (Gate+): {adc6:.5f} V\n"
                    f"Vsd: {vsd:.5f} V, Is: {i_source:.6f} A, Ig: {i_gate:.6f} A, Id: {i_drain:.6f} A, Isd: {isd:.6f} A\n"
                    "-----------------------------------------\n"
                )
            else:
                log_entry = (
                    f"Time: {current_time}\n"
                    f"ADC1 (Src-): {adc1:.5f} V, ADC2 (Drn-): {adc2:.5f} V, ADC3 (Gate-): {adc3:.5f} V\n"
                    f"ADC4 (Src+): {adc4:.5f} V, ADC5 (Drn+): {adc5:.5f} V, ADC6 (Gate+): {adc6:.5f} V\n"
                    f"Vsg: {vsg:.5f} V, Is: {i_source:.6f} A, Ig: {i_gate:.6f} A, Id: {i_drain:.6f} A\n"
                    "-----------------------------------------\n"
                )
            
            output_text.insert(tk.END, log_entry)
            output_text.see(tk.END)

            current_voltage += sweep_step
            steps_taken += 1

            if collecting_data:
                root.after(int(interval * 1000), collect_step)

        collect_step()

    collect_data()


def save_data(chip_name, trial_name, sweep_min, sweep_max, sweep_points):
    base_dir = "/home/pi/Desktop/Biosensor V2/data"
    chip_path = os.path.join(base_dir, chip_name)
    if not os.path.exists(chip_path):
        os.makedirs(chip_path)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    trial_folder_name = f"{trial_name} - {timestamp}"
    trial_path = os.path.join(chip_path, trial_folder_name)
    os.makedirs(trial_path, exist_ok=True)

    if mode == "output":
        save_excel_file(
            trial_path,
            f"{trial_name}_VSD_ISID_{timestamp}.xlsx",
            "OSDL Biosensor V1 - Output",
            sweep_min,
            sweep_max,
            sweep_points,
            current_key="i_drain",
            x_key="vsd",
            y_key="isd",
        )

        save_excel_file(
            trial_path,
            f"{trial_name}_VSD_ISIG_{timestamp}.xlsx",
            "OSDL Biosensor V1 - Output",
            sweep_min,
            sweep_max,
            sweep_points,
            current_key="i_source",
            x_key="vsd",
            y_key="isd",
        )
    else:
        save_excel_file(
            trial_path,
            f"{trial_name}_VSG_IDIS_{timestamp}.xlsx",
            "OSDL Biosensor V1 - Transfer",
            sweep_min,
            sweep_max,
            sweep_points,
            current_key="i_source",
            x_key="vsg",
            y_key="i_drain",
        )


def save_excel_file(
    directory, filename, title, sweep_min, sweep_max, sweep_points, current_key, x_key, y_key
):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data Collection"

    worksheet.cell(row=1, column=1, value=title)
    worksheet.cell(row=2, column=1, value="Sweep Min:")
    worksheet.cell(row=2, column=2, value=sweep_min)
    worksheet.cell(row=3, column=1, value="Sweep Max:")
    worksheet.cell(row=3, column=2, value=sweep_max)
    worksheet.cell(row=4, column=1, value="Sweep Points:")
    worksheet.cell(row=4, column=2, value=sweep_points)

    headers = ["X-Axis (V)"]
    for dac1_voltage in data_by_dac1:
        headers.append(f"{dac1_voltage:.6f}")
    for dac1_voltage in data_by_dac1:
        headers.append(f"{dac1_voltage:.6f}")
    worksheet.append(headers)

    max_length = max(
        len(data_by_dac1[dac1_voltage][x_key]) for dac1_voltage in data_by_dac1
    )
    for i in range(max_length):
        row_data = []

        x_value = (
            data_by_dac1[next(iter(data_by_dac1))][x_key][i]
            if i < len(data_by_dac1[next(iter(data_by_dac1))][x_key])
            else ""
        )
        row_data.append(x_value)

        for dac1_voltage in data_by_dac1:
            if mode == "output":
                y_value = (
                    data_by_dac1[dac1_voltage][y_key][i]
                    if i < len(data_by_dac1[dac1_voltage][y_key])
                    else ""
                )
            else:
                y_value = (
                    data_by_dac1[dac1_voltage][y_key][i]
                    if i < len(data_by_dac1[dac1_voltage][y_key])
                    else ""
                )
            row_data.append(y_value)

        for dac1_voltage in data_by_dac1:
            current_value = (
                data_by_dac1[dac1_voltage][current_key][i]
                if i < len(data_by_dac1[dac1_voltage][current_key])
                else ""
            )
            row_data.append(current_value)

        worksheet.append(row_data)

    file_path = os.path.join(directory, filename)
    workbook.save(file_path)
    print(f"Data saved to {file_path}")


def reset_plot():
    global ax, canvas, mode
    ax.clear()
    if mode == "output":
        ax.set_xlabel("Vsd (V)")
        ax.set_ylabel("Isd (A)")
    else:
        ax.set_xlabel("Vsg (V)")
        ax.set_ylabel("Id (A)")
    ax.set_xlim(-5, 5)
    ax.set_ylim(-1, 1)
    canvas.draw_idle()


def reset_parameters(params_entries):
    global mode
    params_entries["duration"].delete(0, tk.END)
    params_entries["duration"].insert(0, "120")

    params_entries["sweep_min"].delete(0, tk.END)
    params_entries["sweep_min"].insert(0, "3.0")

    params_entries["sweep_max"].delete(0, tk.END)
    params_entries["sweep_max"].insert(0, "0.0")

    params_entries["sweep_step"].delete(0, tk.END)
    params_entries["sweep_step"].insert(0, "-0.1")

    params_entries["constant_voltage"].delete(0, tk.END)
    params_entries["constant_voltage"].insert(0, "0.5")

    params_entries["step_interval"].delete(0, tk.END)
    params_entries["step_interval"].insert(0, "1")


def create_data_setup_frame(parent):
    frame_data_setup = tk.LabelFrame(
        parent, text="Data Setup", relief=tk.SUNKEN, borderwidth=2
    )
    frame_data_setup.pack(fill="x", padx=10, pady=10)

    tk.Label(frame_data_setup, text="Chip Name:").grid(
        row=0, column=0, sticky="e", padx=5, pady=5
    )
    chip_name_entry = tk.Entry(frame_data_setup, width=20)
    chip_name_entry.grid(row=0, column=1, padx=5, pady=5)
    chip_name_entry.insert(0, "Default Chip")

    tk.Label(frame_data_setup, text="Trial Name:").grid(
        row=1, column=0, sticky="e", padx=5, pady=5
    )
    trial_name_entry = tk.Entry(frame_data_setup, width=20)
    trial_name_entry.grid(row=1, column=1, padx=5, pady=5)
    trial_name_entry.insert(0, "Trial 1")

    tk.Label(frame_data_setup, text="Mode:").grid(
        row=2, column=0, sticky="e", padx=5, pady=5
    )
    
    mode_var = tk.StringVar(value="output")
    
    mode_frame = tk.Frame(frame_data_setup)
    mode_frame.grid(row=2, column=1, padx=5, pady=5, sticky="w")
    
    output_button = tk.Radiobutton(mode_frame, text="Output", variable=mode_var, 
                                  value="output", width=8)
    output_button.pack(side=tk.LEFT)
    
    transfer_button = tk.Radiobutton(mode_frame, text="Transfer", variable=mode_var, 
                                    value="transfer", width=8)
    transfer_button.pack(side=tk.LEFT)

    return chip_name_entry, trial_name_entry, mode_var


def create_params_frame(parent):
    frame_params = tk.LabelFrame(
        parent, text="Parameters", relief=tk.SUNKEN, borderwidth=2
    )
    frame_params.pack(fill="x", padx=10, pady=10)

    params_entries = {}

    tk.Label(frame_params, text="Duration (s):").grid(
        row=0, column=0, sticky="e", padx=5, pady=5
    )
    duration_entry = tk.Entry(frame_params, width=20)
    duration_entry.grid(row=0, column=1, columnspan=3, pady=5, sticky="w")
    duration_entry.insert(0, "120")
    params_entries["duration"] = duration_entry

    tk.Label(frame_params, text="Sweep Voltage (V):").grid(
        row=1, column=0, sticky="e", padx=5, pady=5
    )
    sweep_min_entry = tk.Entry(frame_params, width=8)
    sweep_min_entry.grid(row=1, column=1, padx=(0, 2), pady=5, sticky="w")
    sweep_min_entry.insert(0, "3.0")
    params_entries["sweep_min"] = sweep_min_entry
    tk.Label(frame_params, text="to").grid(row=1, column=2, padx=(2, 2), sticky="w")
    sweep_max_entry = tk.Entry(frame_params, width=8)
    sweep_max_entry.grid(row=1, column=3, padx=(2, 5), pady=5, sticky="w")
    sweep_max_entry.insert(0, "0.0")
    params_entries["sweep_max"] = sweep_max_entry

    tk.Label(frame_params, text="Step Voltage (V):").grid(
        row=2, column=0, sticky="e", padx=5, pady=5
    )
    sweep_step_entry = tk.Entry(frame_params, width=20)
    sweep_step_entry.grid(row=2, column=1, columnspan=3, pady=5, sticky="w")
    sweep_step_entry.insert(0, "-0.1")
    params_entries["sweep_step"] = sweep_step_entry

    tk.Label(frame_params, text="Step Interval (s):").grid(
        row=3, column=0, sticky="e", padx=5, pady=5
    )
    step_interval_entry = tk.Entry(frame_params, width=20)
    step_interval_entry.grid(row=3, column=1, columnspan=3, pady=5, sticky="w")
    step_interval_entry.insert(0, "1")
    params_entries["step_interval"] = step_interval_entry

    tk.Label(frame_params, text="Constant Voltage (V):").grid(
        row=4, column=0, sticky="e", padx=5, pady=5
    )
    constant_voltage_entry = tk.Entry(frame_params, width=20)
    constant_voltage_entry.grid(row=4, column=1, columnspan=3, pady=5, sticky="w")
    constant_voltage_entry.insert(0, "0.5")
    params_entries["constant_voltage"] = constant_voltage_entry

    return params_entries


def create_plot_frame(parent):
    global fig, ax, canvas

    fig, ax = plt.subplots(figsize=(5, 3))

    fig.subplots_adjust(left=0.15, right=0.95, top=0.9, bottom=0.15)

    ax.set_xlabel("Vsd (V)")
    ax.set_ylabel("Isd (A)")
    ax.set_xlim(-5, 5)
    ax.set_ylim(-1, 1)

    canvas = FigureCanvasTkAgg(fig, master=parent)
    canvas_widget = canvas.get_tk_widget()

    toolbar = NavigationToolbar2Tk(canvas, parent)
    toolbar.update()

    toolbar.pack(side=tk.BOTTOM, fill=tk.X, expand=False)

    canvas_widget.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

    canvas.draw()


def update_plot(constant_voltage, x_value, y_value, x_label, y_label):
    global fig, ax, canvas, plot_colors, data_by_dac1, mode

    ax.clear()

    if mode == "output":
        for i, const_v in enumerate(data_by_dac1):
            color = plot_colors[i % len(plot_colors)]
            ax.plot(
                data_by_dac1[const_v]["vsd"],
                data_by_dac1[const_v]["isd"],
                label=f"Gate = {const_v:.2f} V",
                color=color,
            )
    else:
        for i, const_v in enumerate(data_by_dac1):
            color = plot_colors[i % len(plot_colors)]
            ax.plot(
                data_by_dac1[const_v]["vsg"],
                data_by_dac1[const_v]["i_drain"],
                label=f"Drain = {const_v:.2f} V",
                color=color,
            )

    ax.set_xlabel(x_label)
    ax.set_ylabel(y_label)
    ax.legend(loc="upper left")
    ax.relim()
    ax.autoscale_view()

    canvas.draw_idle()


def create_console_log_frame(parent):
    frame_output = tk.LabelFrame(
        parent, text="Console Log", relief=tk.SUNKEN, borderwidth=2
    )
    frame_output.pack(fill="both", padx=10, pady=0, expand=True)

    global output_text
    output_text = tk.Text(frame_output, height=6, width=40, wrap="none", bg="white")
    scrollbar = tk.Scrollbar(
        frame_output, orient=tk.VERTICAL, command=output_text.yview
    )
    output_text.config(yscrollcommand=scrollbar.set)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    output_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    return output_text